"""
Extraction Automatique de Factures - Version finale
"""

import streamlit as st
import anthropic
import base64
import json
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import datetime
from typing import List, Dict, Optional
import re

# ── Page config ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Extraction Factures",
    page_icon="🧾",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# ── CSS ───────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
    .main-header {
        background: linear-gradient(135deg, #1e3a8a, #3b82f6);
        color: white; padding: 20px 30px; border-radius: 12px;
        margin-bottom: 25px; text-align: center;
        box-shadow: 0 4px 6px -1px rgba(0,0,0,0.1);
    }
    .main-header h1 { margin: 0; font-size: 2rem; font-weight: 600; }
    .main-header p  { margin: 5px 0 0; opacity: 0.85; font-size: 1rem; }
    .info-box {
        background: #eff6ff; border-left: 4px solid #3b82f6;
        padding: 12px 16px; border-radius: 6px; margin-bottom: 15px;
        font-size: 0.9rem; color: #1e3a8a;
    }
    .result-box {
        background: #f0fdf4; border-left: 4px solid #22c55e;
        padding: 14px 18px; border-radius: 8px; margin: 8px 0; color: #166534;
    }
    .error-box {
        background: #fef2f2; border-left: 4px solid #ef4444;
        padding: 14px 18px; border-radius: 8px; margin: 8px 0; color: #991b1b;
    }
    .stButton > button {
        background: linear-gradient(135deg, #1e3a8a, #3b82f6);
        color: white; border: none; border-radius: 8px;
        padding: 10px 24px; font-size: 1rem; font-weight: 600;
        width: 100%; transition: all .2s;
    }
    .stButton > button:hover { opacity: .85; transform: translateY(-1px); }
    .footer {
        text-align: center; margin-top: 40px; padding: 20px;
        color: #6b7280; font-size: 0.875rem; border-top: 1px solid #e5e7eb;
    }
</style>
""", unsafe_allow_html=True)

# ── Header ────────────────────────────────────────────────────────────────────
st.markdown("""
<div class="main-header">
    <h1>🧾 Extraction Automatique de Factures</h1>
    <p>Importez vos factures PDF/image et exportez les données structurées vers Excel</p>
</div>
""", unsafe_allow_html=True)

# ── Session state ─────────────────────────────────────────────────────────────
for key, val in [("extracted_data", []), ("api_key_configured", False)]:
    if key not in st.session_state:
        st.session_state[key] = val

# ── Clé API ───────────────────────────────────────────────────────────────────
def get_api_key() -> Optional[str]:
    try:
        return st.secrets["ANTHROPIC_API_KEY"]
    except Exception:
        return st.session_state.get("manual_api_key")

# ── Sidebar ───────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("### ⚙️ Configuration")
    st.markdown("---")
    key_from_secrets = False
    try:
        _ = st.secrets["ANTHROPIC_API_KEY"]
        key_from_secrets = True
    except Exception:
        pass

    if key_from_secrets:
        st.success("✅ Clé API chargée depuis les secrets")
        st.session_state.api_key_configured = True
    else:
        api_key_input = st.text_input("Clé API Anthropic", type="password", placeholder="sk-ant-...")
        if api_key_input:
            st.session_state.manual_api_key = api_key_input
            st.session_state.api_key_configured = True
            st.success("✅ Clé API configurée")
        else:
            st.session_state.api_key_configured = False
            st.warning("⚠️ Clé API requise")

    st.markdown("---")
    st.markdown("""### 📋 Instructions
1. Configurez votre clé API Anthropic
2. Importez vos factures (PDF, PNG, JPG)
3. Cliquez sur **Extraire**
4. Téléchargez le fichier Excel

**Formats :** PDF · PNG · JPG/JPEG""")
    st.markdown("---")
    st.markdown(f"**v2.0** | {datetime.now().strftime('%d/%m/%Y')}")

# ── Prompt extraction ─────────────────────────────────────────────────────────
PROMPT = """Tu es un expert comptable. Analyse cette facture/document et extrais les informations.

Réponds UNIQUEMENT avec ce JSON (rien avant, rien après) :
{
  "fournisseur": "raison sociale de l'entreprise qui ÉMET la facture",
  "date": "date de la facture format JJ/MM/AAAA",
  "commande": "numéro bon de commande (BC, PO, CMDE, N°BC, Order...)",
  "bon_de_livraison": "numéro bon de livraison (BL, N°BL, Livraison Réf...)",
  "numero_facture": "numéro de facture (FAC, FACT, N°, Invoice No...)",
  "montant_facture": "montant TOTAL TTC en chiffres sans espace ni symbole (ex: 185195.40)"
}

RÈGLES :
- fournisseur = entité en haut du document qui émet/vend
- montant = TOTAL TTC final (après TVA, le plus grand montant)
- champ absent = null
- JSON UNIQUEMENT, rien d'autre"""

# ── Helpers ───────────────────────────────────────────────────────────────────
def encode_b64(data: bytes) -> str:
    return base64.standard_b64encode(data).decode("utf-8")

def parse_json_response(raw: str) -> dict:
    clean = re.sub(r"^```(?:json)?\s*", "", raw.strip())
    clean = re.sub(r"\s*```$", "", clean).strip()
    clean = re.sub(r'\s+', ' ', clean)
    try:
        return json.loads(clean)
    except json.JSONDecodeError:
        m = re.search(r'\{.*\}', clean, re.DOTALL)
        if m:
            return json.loads(m.group())
        raise ValueError(f"JSON invalide: {clean[:200]}")

def normalize_fields(data: dict, filename: str) -> dict:
    for f in ["fournisseur", "date", "commande", "bon_de_livraison", "numero_facture", "montant_facture"]:
        if data.get(f) in [None, "null", "NULL", "None", "", "N/A", "n/a"]:
            data[f] = None
    data["fichier"] = filename
    return data

def extract_invoice_info(file_bytes: bytes, mime: str, filename: str) -> Dict:
    """Extraction principale avec 3 stratégies."""
    api_key = get_api_key()
    if not api_key:
        raise ValueError("Clé API non configurée.")

    client = anthropic.Anthropic(api_key=api_key)
    last_error = None

    # ══ Stratégie 1 : PDF natif avec beta header (scans inclus) ══════════════
    if mime == "application/pdf":
        try:
            resp = client.beta.messages.create(
                model="claude-opus-4-5",
                max_tokens=1024,
                betas=["pdfs-2024-09-25"],
                messages=[{"role": "user", "content": [
                    {"type": "document", "source": {
                        "type": "base64",
                        "media_type": "application/pdf",
                        "data": encode_b64(file_bytes)
                    }},
                    {"type": "text", "text": PROMPT}
                ]}]
            )
            data = parse_json_response(resp.content[0].text)
            # Vérifier si on a au moins 1 champ non-null
            fields = ["fournisseur", "date", "commande", "bon_de_livraison", "numero_facture", "montant_facture"]
            has_data = any(data.get(f) not in [None, "null", "NULL", "None", ""] for f in fields)
            if has_data:
                return normalize_fields(data, filename)
        except Exception as e:
            last_error = str(e)

    # ══ Stratégie 2 : PDF natif avec claude-3-5-sonnet standard ══════════════
    if mime == "application/pdf":
        try:
            resp = client.messages.create(
                model="claude-3-5-sonnet-20241022",
                max_tokens=1024,
                temperature=0,
                messages=[{"role": "user", "content": [
                    {"type": "document", "source": {
                        "type": "base64",
                        "media_type": "application/pdf",
                        "data": encode_b64(file_bytes)
                    }},
                    {"type": "text", "text": PROMPT}
                ]}]
            )
            data = parse_json_response(resp.content[0].text)
            fields = ["fournisseur", "date", "commande", "bon_de_livraison", "numero_facture", "montant_facture"]
            has_data = any(data.get(f) not in [None, "null", "NULL", "None", ""] for f in fields)
            if has_data:
                return normalize_fields(data, filename)
        except Exception as e:
            last_error = str(e)

    # ══ Stratégie 3 : Conversion PDF → images via PyMuPDF ════════════════════
    if mime == "application/pdf":
        try:
            import fitz
            doc = fitz.open(stream=file_bytes, filetype="pdf")
            img_blocks = []
            for i in range(min(2, len(doc))):
                pix = doc[i].get_pixmap(matrix=fitz.Matrix(2.0, 2.0))
                img_b64 = base64.standard_b64encode(pix.tobytes("jpeg")).decode()
                img_blocks.append({
                    "type": "image",
                    "source": {"type": "base64", "media_type": "image/jpeg", "data": img_b64}
                })
            doc.close()
            img_blocks.append({"type": "text", "text": PROMPT})
            resp = client.messages.create(
                model="claude-3-5-sonnet-20241022",
                max_tokens=1024,
                temperature=0,
                messages=[{"role": "user", "content": img_blocks}]
            )
            data = parse_json_response(resp.content[0].text)
            return normalize_fields(data, filename)
        except Exception as e:
            last_error = str(e)

    # ══ Stratégie 4 : Image directe (PNG/JPG) ════════════════════════════════
    if mime != "application/pdf":
        try:
            resp = client.messages.create(
                model="claude-3-5-sonnet-20241022",
                max_tokens=1024,
                temperature=0,
                messages=[{"role": "user", "content": [
                    {"type": "image", "source": {
                        "type": "base64", "media_type": mime,
                        "data": encode_b64(file_bytes)
                    }},
                    {"type": "text", "text": PROMPT}
                ]}]
            )
            data = parse_json_response(resp.content[0].text)
            return normalize_fields(data, filename)
        except Exception as e:
            last_error = str(e)

    # Si toutes les stratégies échouent
    raise ValueError(f"Toutes les stratégies ont échoué. Dernière erreur : {last_error}")

def parse_montant(v) -> float:
    if v in (None, "null", ""):
        return 0.0
    try:
        if isinstance(v, str):
            v = re.sub(r'[^\d.,-]', '', v).replace(',', '.')
            if v.count('.') > 1:
                v = v.replace('.', '', v.count('.') - 1)
        return float(v)
    except Exception:
        return 0.0

def build_excel(records: List[Dict]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Factures"

    hfill  = PatternFill("solid", fgColor="1e3a8a")
    afill  = PatternFill("solid", fgColor="dbeafe")
    wfill  = PatternFill("solid", fgColor="FFFFFF")
    border = Border(*[Side(style="thin", color="d1d5db")]*4,
                    **{s: Side(style="thin", color="d1d5db")
                       for s in ["left","right","top","bottom"]})
    border = Border(
        left=Side(style="thin", color="d1d5db"),
        right=Side(style="thin", color="d1d5db"),
        top=Side(style="thin", color="d1d5db"),
        bottom=Side(style="thin", color="d1d5db"),
    )

    cols = [
        ("Fournisseur",      "fournisseur",      25),
        ("Date",             "date",             14),
        ("N° Commande",      "commande",         18),
        ("Bon de Livraison", "bon_de_livraison", 20),
        ("N° Facture",       "numero_facture",   20),
        ("Montant TTC",      "montant_facture",  16),
        ("Fichier Source",   "fichier",          30),
    ]

    for ci, (label, _, w) in enumerate(cols, 1):
        c = ws.cell(row=1, column=ci, value=label)
        c.font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
        c.fill = hfill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 30

    total = 0.0
    for ri, rec in enumerate(records, 2):
        fill = afill if ri % 2 == 0 else wfill
        for ci, (_, key, _) in enumerate(cols, 1):
            val = rec.get(key)
            if key == "montant_facture":
                val = parse_montant(val)
                total += val
            c = ws.cell(row=ri, column=ci, value=val if val not in (None, "null", "") else "")
            c.font = Font(name="Arial", size=10)
            c.fill = fill
            c.alignment = Alignment(horizontal="right" if key == "montant_facture" else "left", vertical="center")
            c.border = border
            if key == "montant_facture" and isinstance(val, float):
                c.number_format = '#,##0.00 "MAD"'
        ws.row_dimensions[ri].height = 20

    lr = len(records) + 2
    ws.merge_cells(start_row=lr, start_column=1, end_row=lr, end_column=5)
    lbl = ws.cell(row=lr, column=1, value="TOTAL GÉNÉRAL")
    lbl.font = Font(bold=True, name="Arial", size=11)
    lbl.fill = PatternFill("solid", fgColor="fbbf24")
    lbl.border = border
    tot = ws.cell(row=lr, column=6, value=total)
    tot.font = Font(bold=True, name="Arial", size=11)
    tot.number_format = '#,##0.00 "MAD"'
    tot.alignment = Alignment(horizontal="right")
    tot.fill = PatternFill("solid", fgColor="fbbf24")
    tot.border = border

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

# ── Interface ─────────────────────────────────────────────────────────────────
col_left, col_right = st.columns([1, 2])

with col_left:
    st.markdown("### 📤 Importer des Factures")
    st.markdown("""<div class="info-box">
        <strong>Formats acceptés :</strong> PDF, PNG, JPG, JPEG<br>
        <strong>Taille max :</strong> 200 MB par fichier
    </div>""", unsafe_allow_html=True)

    uploaded_files = st.file_uploader(
        "Choisissez vos fichiers",
        type=["pdf", "png", "jpg", "jpeg"],
        accept_multiple_files=True
    )

    if uploaded_files:
        st.markdown(f'<div class="result-box">✅ {len(uploaded_files)} fichier(s) sélectionné(s)</div>',
                    unsafe_allow_html=True)
        with st.expander("📋 Fichiers sélectionnés"):
            for f in uploaded_files:
                st.text(f"📄 {f.name} ({f.size/1024:.1f} KB)")

    extract_btn = st.button(
        "🔍 Extraire les Informations",
        disabled=not (uploaded_files and st.session_state.api_key_configured),
        use_container_width=True
    )

    if st.session_state.extracted_data:
        st.markdown("---")
        st.markdown("### 💾 Exporter")
        excel_bytes = build_excel(st.session_state.extracted_data)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        c1, c2 = st.columns(2)
        with c1:
            st.download_button(
                "📥 Télécharger Excel", data=excel_bytes,
                file_name=f"factures_{ts}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
        with c2:
            if st.button("🗑️ Effacer", use_container_width=True):
                st.session_state.extracted_data = []

with col_right:
    st.markdown("### 📊 Données Extraites")

    if extract_btn and uploaded_files:
        mime_map = {"pdf": "application/pdf", "png": "image/png",
                    "jpg": "image/jpeg", "jpeg": "image/jpeg"}
        progress = st.progress(0, text="Démarrage…")
        newly = []

        for i, uf in enumerate(uploaded_files):
            ext = uf.name.rsplit(".", 1)[-1].lower()
            mime = mime_map.get(ext, "application/octet-stream")
            progress.progress(i / len(uploaded_files),
                              text=f"⏳ {uf.name} ({i+1}/{len(uploaded_files)})…")
            try:
                info = extract_invoice_info(uf.read(), mime, uf.name)
                newly.append(info)
                # Afficher résultat immédiatement
                fields_found = [k for k in ["fournisseur","date","numero_facture","montant_facture"]
                                if info.get(k) not in (None, "null", "")]
                if fields_found:
                    st.markdown(f"""<div class="result-box">
                        ✅ <strong>{uf.name}</strong><br>
                        🏢 {info.get('fournisseur') or '—'} &nbsp;|&nbsp;
                        📅 {info.get('date') or '—'} &nbsp;|&nbsp;
                        🔢 {info.get('numero_facture') or '—'} &nbsp;|&nbsp;
                        💰 {info.get('montant_facture') or '—'} MAD
                    </div>""", unsafe_allow_html=True)
                else:
                    st.markdown(f'<div class="error-box">⚠️ {uf.name} — aucune donnée extraite</div>',
                                unsafe_allow_html=True)
            except Exception as e:
                err_info = {"fournisseur": None, "date": None, "commande": None,
                            "bon_de_livraison": None, "numero_facture": None,
                            "montant_facture": None, "fichier": uf.name, "erreur": str(e)}
                newly.append(err_info)
                st.markdown(f'<div class="error-box">❌ {uf.name} — {e}</div>',
                            unsafe_allow_html=True)
            progress.progress((i+1) / len(uploaded_files), text=f"({i+1}/{len(uploaded_files)}) fait")

        st.session_state.extracted_data.extend(newly)
        progress.empty()
        ok = sum(1 for r in newly if r.get("fournisseur") or r.get("numero_facture"))
        st.success(f"✅ Terminé — {ok}/{len(newly)} facture(s) extraite(s) avec succès")

    if st.session_state.extracted_data:
        st.markdown("---")
        df = pd.DataFrame(st.session_state.extracted_data)
        col_map = {
            'fournisseur': 'Fournisseur', 'date': 'Date',
            'commande': 'N° Commande', 'bon_de_livraison': 'Bon Livraison',
            'numero_facture': 'N° Facture', 'montant_facture': 'Montant (MAD)',
            'fichier': 'Fichier'
        }
        df2 = df.rename(columns=col_map)
        show = [c for c in ['Fournisseur','Date','N° Commande','Bon Livraison',
                            'N° Facture','Montant (MAD)','Fichier'] if c in df2.columns]
        st.dataframe(df2[show], use_container_width=True, height=min(400, 60+len(df2)*35), hide_index=True)

        st.markdown("---")
        c1, c2, c3, c4 = st.columns(4)
        total = sum(parse_montant(r.get("montant_facture")) for r in st.session_state.extracted_data)
        valid = sum(1 for r in st.session_state.extracted_data if parse_montant(r.get("montant_facture")) > 0)
        c1.metric("📄 Traitées",  len(st.session_state.extracted_data))
        c2.metric("💰 Total",     f"{total:,.2f} MAD")
        c3.metric("✅ Réussies",  f"{valid}/{len(st.session_state.extracted_data)}")
        c4.metric("📅 Date",      datetime.now().strftime("%d/%m %H:%M"))
    else:
        st.markdown("""
        <div style="text-align:center;padding:60px 20px;background:#f8fafc;border-radius:8px;">
            <div style="font-size:4rem;margin-bottom:15px;">🧾</div>
            <p style="font-size:1.1rem;color:#1e293b;">Aucune donnée extraite</p>
            <p style="color:#64748b;">Importez une facture à gauche et cliquez sur <strong>Extraire</strong></p>
        </div>""", unsafe_allow_html=True)

st.markdown("""
<div class="footer">
    🧾 Extraction Automatique de Factures | Claude API + Streamlit
</div>""", unsafe_allow_html=True)
